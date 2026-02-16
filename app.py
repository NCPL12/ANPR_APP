"""
ANPR REST API – Python backend for MongoDB detected_plates.
Serves stats and paginated plate records for the dashboard.
"""
import io
import os
from datetime import datetime, timedelta, timezone

from dotenv import load_dotenv
load_dotenv()
from bson import ObjectId
import base64
from flask import Flask, Response, jsonify, request, send_from_directory, send_file
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from flask_cors import CORS
from pymongo import MongoClient, DESCENDING, ASCENDING

app = Flask(__name__, static_folder=".", static_url_path="")
CORS(app)

# MongoDB connection
MONGO_URI = os.environ.get(
    "MONGODB_URI",
    "mongodb+srv://anprncpl_db_user:aYWS8VfcQMagPrX0@anpr.w8cukh3.mongodb.net/"
)
DB_NAME = os.environ.get("ANPR_DB_NAME", "anpr_database")
COLLECTION_NAME = "detected_plates"

client = MongoClient(MONGO_URI)
db = client[DB_NAME]
coll = db[COLLECTION_NAME]

# Map vehicle_class to plate type (adjust as per your codes)
VEHICLE_CLASS_LABELS = {
    0: "Unknown",
    1: "Private",
    2: "Commercial",
    3: "Government",
    4: "Special",
    5: "Diplomatic",
    6: "Taxi",
    7: "Truck",
}


def json_serial(obj):
    """Convert ObjectId and datetime for JSON."""
    if isinstance(obj, ObjectId):
        return str(obj)
    if isinstance(obj, datetime):
        return obj.isoformat()
    raise TypeError(f"Object of type {type(obj)} is not JSON serializable")


def to_iso_utc(dt):
    """Return ISO 8601 string with Z (UTC) so frontend parses correctly."""
    if dt is None:
        return None
    if not isinstance(dt, datetime):
        return None
    if dt.tzinfo is not None:
        dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
    return dt.strftime("%Y-%m-%dT%H:%M:%S.%f")[:-3] + "Z"


def doc_to_plate(doc, include_image=False):
    """Convert MongoDB document to API response shape (all fields from DB)."""
    ts = doc.get("timestamp") or doc.get("created_at")
    created = doc.get("created_at")
    plate = {
        "id": str(doc["_id"]),
        "plate_number": doc.get("plate_number", ""),
        "raw_text": doc.get("raw_text", ""),
        "confidence": doc.get("confidence"),
        "ocr_engine": doc.get("ocr_engine", ""),
        "timestamp": to_iso_utc(ts),
        "frame_coords": doc.get("frame_coords"),
        "vehicle_coords": doc.get("vehicle_coords"),
        "vehicle_confidence": doc.get("vehicle_confidence"),
        "vehicle_class": doc.get("vehicle_class"),
        "plate_type": VEHICLE_CLASS_LABELS.get(doc.get("vehicle_class"), "Unknown"),
        "image_saved": doc.get("image_saved", False),
        "created_at": to_iso_utc(created),
    }
    if include_image and doc.get("plate_image"):
        plate["plate_image"] = doc["plate_image"]
    return plate


@app.route("/")
def index():
    return send_from_directory(".", "anpr.html")


@app.route("/api/stats", methods=["GET"])
def get_stats():
    """Dashboard stats: total, today, this week. Uses timestamp or created_at. Cameras/sites from env."""
    now = datetime.utcnow()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    week_start = today_start - timedelta(days=today_start.weekday())

    total = coll.count_documents({})

    # Use timestamp or created_at (whichever exists) and count only up to now
    def count_in_range(start, end):
        pipeline = [
            {
                "$match": {
                    "$expr": {
                        "$and": [
                            {"$gte": [{"$ifNull": ["$timestamp", "$created_at"]}, {"$literal": start}]},
                            {"$lte": [{"$ifNull": ["$timestamp", "$created_at"]}, {"$literal": end}]},
                        ]
                    }
                }
            },
            {"$count": "n"},
        ]
        result = list(coll.aggregate(pipeline))
        return result[0]["n"] if result else 0

    today_count = count_in_range(today_start, now)
    week_count = count_in_range(week_start, now)

    return jsonify({
        "total": total,
        "today": today_count,
        "week": week_count,
        "cameras": int(os.environ.get("ANPR_CAMERAS", "23")),
        "sites": int(os.environ.get("ANPR_SITES", "2")),
    })


@app.route("/api/plates", methods=["GET"])
def get_plates():
    """List detected plates with pagination and sort. Use limit=all or high limit to get all records."""
    page = max(1, int(request.args.get("page", 1)))
    limit_arg = request.args.get("limit", "25")
    if limit_arg.lower() == "all":
        limit = 50000
    else:
        limit = min(50000, max(1, int(limit_arg) if limit_arg.isdigit() else 25))
    sort_param = request.args.get("sort", "date-desc")

    sort_map = {
        "date-desc": [("timestamp", DESCENDING), ("created_at", DESCENDING)],
        "date-asc": [("timestamp", ASCENDING), ("created_at", ASCENDING)],
        "plate": [("plate_number", ASCENDING)],
        "site": [("ocr_engine", ASCENDING), ("timestamp", DESCENDING), ("created_at", DESCENDING)],
    }
    sort = sort_map.get(sort_param, sort_map["date-desc"])

    skip = (page - 1) * limit
    cursor = coll.find({}).sort(sort).skip(skip).limit(limit)
    total = coll.count_documents({})

    items = [doc_to_plate(d, include_image=False) for d in cursor]
    return jsonify({
        "items": items,
        "total": total,
        "page": page,
        "limit": limit,
        "pages": (total + limit - 1) // limit if limit else 0,
    })


@app.route("/api/plates/<plate_id>", methods=["GET"])
def get_plate(plate_id):
    """Single plate by id, optionally with base64 image."""
    try:
        doc = coll.find_one({"_id": ObjectId(plate_id)})
    except Exception:
        return jsonify({"error": "Invalid id"}), 400
    if not doc:
        return jsonify({"error": "Not found"}), 404
    return jsonify(doc_to_plate(doc, include_image=True))


@app.route("/api/plates/<plate_id>/image", methods=["GET"])
def get_plate_image(plate_id):
    """Return plate image as JSON { image: base64 } or 404."""
    try:
        doc = coll.find_one({"_id": ObjectId(plate_id)}, {"plate_image": 1})
    except Exception:
        return jsonify({"error": "Invalid id"}), 400
    if not doc or not doc.get("plate_image"):
        return jsonify({"error": "Image not found"}), 404
    return jsonify({"image": doc["plate_image"]})


@app.route("/api/plates/<plate_id>/image/img", methods=["GET"])
def get_plate_image_binary(plate_id):
    """Return plate image as JPEG for <img> src."""
    try:
        doc = coll.find_one({"_id": ObjectId(plate_id)}, {"plate_image": 1})
    except Exception:
        return Response(status=400)
    if not doc or not doc.get("plate_image"):
        return Response(status=404)
    try:
        raw = base64.b64decode(doc["plate_image"])
    except Exception:
        return Response(status=500)
    return Response(raw, mimetype="image/jpeg")


# Target size for embedded images in Excel (width, height) in pixels
EXCEL_IMAGE_WIDTH = 80
EXCEL_IMAGE_HEIGHT = 44


@app.route("/api/export/excel", methods=["GET"])
def export_excel():
    """Export plates in date range to Excel (.xlsx) with images."""
    from_arg = request.args.get("from", "").strip()
    to_arg = request.args.get("to", "").strip()
    try:
        from_dt = datetime.fromisoformat(from_arg.replace("Z", "+00:00")) if from_arg else None
    except (ValueError, TypeError):
        from_dt = None
    try:
        to_dt = datetime.fromisoformat(to_arg.replace("Z", "+00:00")) if to_arg else None
    except (ValueError, TypeError):
        to_dt = None
    if from_dt and from_dt.tzinfo:
        from_dt = from_dt.astimezone(timezone.utc).replace(tzinfo=None)
    if to_dt and to_dt.tzinfo:
        to_dt = to_dt.astimezone(timezone.utc).replace(tzinfo=None)

    match = {}
    if from_dt is not None or to_dt is not None:
        expr = []
        if from_dt is not None:
            expr.append({"$gte": [{"$ifNull": ["$timestamp", "$created_at"]}, from_dt]})
        if to_dt is not None:
            expr.append({"$lte": [{"$ifNull": ["$timestamp", "$created_at"]}, to_dt]})
        match = {"$expr": {"$and": expr}}

    cursor = (
        coll.find(match)
        .sort([("timestamp", DESCENDING), ("created_at", DESCENDING)])
        .limit(5000)
    )
    docs = list(cursor)

    wb = Workbook()
    ws = wb.active
    ws.title = "ANPR Export"
    headers = ["S.no", "Plate No.", "Plate type", "Date", "Time stamp", "Image"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    ws.row_dimensions[1].height = 20

    for row_idx, doc in enumerate(docs, start=2):
        ts = doc.get("timestamp") or doc.get("created_at")
        date_str = to_iso_utc(ts)
        if date_str:
            try:
                d = datetime.fromisoformat(date_str.replace("Z", "+00:00"))
                date_only = d.strftime("%d.%m.%y")
                time_only = d.strftime("%H:%M:%S")
            except (ValueError, TypeError):
                date_only = date_str[:10] if len(date_str) >= 10 else date_str
                time_only = ""
        else:
            date_only = ""
            time_only = ""

        plate_number = doc.get("plate_number") or ""
        plate_type = VEHICLE_CLASS_LABELS.get(doc.get("vehicle_class"), "Unknown")
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=plate_number)
        ws.cell(row=row_idx, column=3, value=plate_type)
        ws.cell(row=row_idx, column=4, value=date_only)
        ws.cell(row=row_idx, column=5, value=time_only)
        ws.row_dimensions[row_idx].height = max(35, EXCEL_IMAGE_HEIGHT * 0.75)

        img_b64 = doc.get("plate_image")
        if img_b64:
            try:
                raw = base64.b64decode(img_b64)
                img = XLImage(io.BytesIO(raw))
                w, h = img.width, img.height
                if w > EXCEL_IMAGE_WIDTH or h > EXCEL_IMAGE_HEIGHT:
                    try:
                        from PIL import Image as PILImage
                        pil_img = PILImage.open(io.BytesIO(raw))
                        pil_img.thumbnail((EXCEL_IMAGE_WIDTH, EXCEL_IMAGE_HEIGHT))
                        out = io.BytesIO()
                        pil_img.save(out, format="JPEG")
                        out.seek(0)
                        img = XLImage(out)
                    except Exception:
                        pass
                img.anchor = f"F{row_idx}"
                ws.add_image(img)
            except Exception:
                ws.cell(row=row_idx, column=6, value="[image]")

    ws.column_dimensions["F"].width = 14
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    from_label = from_arg[:10] if from_arg else "all"
    to_label = to_arg[:10] if to_arg else "all"
    filename = f"anpr_export_{from_label}_to_{to_label}.xlsx"
    return send_file(
        out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
